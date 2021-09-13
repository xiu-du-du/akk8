from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
wb = load_workbook('gouzi.xlsx')
sh = wb.active

index =[]
tmp = []

for i,cell in enumerate(sh['A']):
    flag = cell.value not in tmp
    if flag:
        tmp.append(cell.value)
    else:
        index.append(i)
fill = PatternFill('solid','FAEEEE')
for i,row in enumerate(sh.rows):
    if i in index:
        for cell in row:
            cell.fill = fill
        print(f'第{i}是重复数据')

wb.save('重复数据2.xlsx')

wb_new = Workbook()
sh_new = wb_new.active
tmp_list = []
for i,row in enumerate(sh.rows):
    if i not in index:
        tmp =[]
        for cell in row:
            tmp.append(cell.value)
        tmp_list.append(tmp)
for row in tmp_list:
    sh_new.append(row)

wb_new.save('非重数据5.xlsx')