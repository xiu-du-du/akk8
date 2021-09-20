# 当前时间:2021/9/20 12:24
from openpyxl import Workbook, load_workbook
from datetime import date


def creat_data():
    wb = Workbook()
    sh = wb.active

    rows = [
        ['Date', 'name', 'time'],
        [date(2050, 12, 1), '小明', '18:50'],
        [date(2050, 12, 2), '小张', '18:10'],
        [date(2050, 12, 3), '小王', '18:20'],
        [date(2050, 12, 4), '小李', '19:52'],
        [date(2050, 12, 5), '小刘', '18:04'],
        [date(2050, 12, 6), '小明', '19:08'],
    ]
    for row in rows:
        sh.append(row)
    wb.save('打卡时间.xlsx')


def statistics():
    wb = load_workbook('打卡时间.xlsx')
    sh = wb['Sheet']
    data=[]
    for i in range(2,sh.max_row+1):
        tmp_list=[]
        for j in range(1,sh.max_column+1):
            tmp_list.append(sh.cell(i,j).value)

        h,m=tmp_list[2].split(':')
        full=int(h)*60+int(m)
        res=full-18*60
        tmp_list.append(res)
        tmp_list[0]=tmp_list[0].date()
        data.append(tmp_list)
    new_wb=Workbook()
    new_sh=new_wb.active
    for row in data:
        new_sh.append(row)
    new_wb.save('加班时间.xlsx')
    new_wb.close()


if __name__ == '__main__':
    creat_data()
    statistics()

