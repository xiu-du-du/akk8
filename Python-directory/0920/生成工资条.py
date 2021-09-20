# 当前时间:2021/9/20 8:51
"""
工号	姓名	部门	基本工资	提成	加班工资	社保扣除	考勤扣除	应发工资	邮箱
10000	刘备	首领	8000	5000	3000	680	0	15320	2463308417@qq.com
10001	关羽	将军	6000	4000	3000	680	0	12320	2463308417@qq.com
10002	张飞	将军	6000	3500	5000	680	300	13520	2463308417@qq.com
10003	诸葛亮	军师	7000	3000	1000	680	600	9720	2463308417@qq.com
10004	狗子	小兵	3000	500	1500	680	0	4320	2463308417@qq.com
10005	蛋蛋	小兵	3000	400	1500	680	0	4220	2463308417@qq.com

"""
from openpyxl import load_workbook
from openpyxl import Workbook
wb=load_workbook('工资数据.xlsx')
sh=wb['Sheet1']
title=[]
# ===========方法一
for i,row in enumerate(sh.rows):
    tmp_list=[]
    for cell in row:
        tmp_list.append(cell.value)
        print(tmp_list)
    if i==0:
        title=tmp_list
    else:
        tmp_wb=Workbook()
        tmp_sh=tmp_wb.active
        tmp_sh.append(title)
        tmp_sh.append(tmp_list)
        tmp_wb.save(f'{i}_{tmp_list[1]}的工资条.xlsx')

# ===========方法二
# count=0
# for row in sh.rows:
#     tmp_list=[]
#     for cell in row:
#         tmp_list.append(cell.value)
#     count+=1
#     if count == 1:
#         title=tmp_list
# print(title)








