# 当前时间： 2021-09-06 22:06
def new():
    from openpyxl import Workbook
    wb = Workbook()    #创建文件对象
    # grab the active worksheet
    sh1 = wb.active     #获取第一个sheet
    sh2 = wb.create_sheet('番剧信息')    # 创建新工作簿，给名称
    wb.save('test2.xlsx')   # 保存文件，给文件名

def set_value():
    from openpyxl import Workbook
    wb=Workbook()
    sh=wb.active
    sh['A1']='hello'   # 写入 hello 数据到 A1 位置的表格
    sh['A2']='python'  # 写入 hello 数据到 A2 位置的表格
    sh['B1']='Excel'   # 写入 hello 数据到 B1 位置的表格
    wb.save('test2.xlsx')

def set_many_value():
    from openpyxl import Workbook
    wb = Workbook()
    sh = wb.active

    rows=[
        ['一人之下',9.8,'冯宝宝'],
        ['黑色五叶草',9.5,'五条悟'],
        ['白色八饼',9.1,'魔法帝']
    ]

    for row in rows:  #读取数据
        sh.append(row)  #数据写入创建的文件
        print(row)   # 这个for循环，row返回的是数据，不是下标
    wb.save('test3.xlsx')
# new()
# set_value()
set_many_value()

















