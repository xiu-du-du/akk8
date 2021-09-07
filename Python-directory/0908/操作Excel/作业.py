# 当前时间： 2021-09-06 22:38

def creat_video():
    from openpyxl import Workbook
    wb=Workbook()
    sh=wb.active

    video_info=[
        ['仙古狂涛',4.0,'俏如来','中国'],
        ['战血天道',9.2,'俏如来','中国'],
        ['齐神录',9.4,'俏如来','中国'],
        ['鬼途奇行录',9.5,'俏如来','中国'],
        ['魅妖记',9.6,'俏如来','中国']
    ]

    for row in video_info:
        sh.append(row)
    wb.save('video.xlsx')

def level():
    from openpyxl import load_workbook
    wb=load_workbook('video.xlsx')
    sh=wb.active
    rows=sh.rows
    a=0
    b=0
    for row in rows:
        if float(row[1].value)>=9.0 :
            a +=1
        if row[3].value=='中国':
            b+=1
    sh2=wb.create_sheet('统计信息')
    sh2.append(['评分大于9',str(a)+'个'])
    sh2.append(['国产',str(b)+'个'])
    wb.save('video.xlsx')
creat_video()
level()