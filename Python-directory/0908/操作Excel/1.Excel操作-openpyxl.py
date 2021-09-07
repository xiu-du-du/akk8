# 当前时间： 2021-09-06 21:09
def open():
    from openpyxl import load_workbook
    wb = load_workbook('test.xlsx') #加载Excel工作表
    #sh1 = wb.active
    sh2 = wb["Sheet1"] #取工作簿
    #sh3 = wb.get_sheet_by_name('Sheet1') #已弃用了 不要用！！
    print( sh2 )

def show_sheets():
    from openpyxl import load_workbook
    #第一种方法：直接获取xlsx文件里的工作簿
    wb = load_workbook('test.xlsx')
    print(wb.sheetnames)
    # 第二种方法：遍历获取xlsx文件里的工作簿
    for sh in wb:
        print(sh.title)

def get_one_value():
    from openpyxl import load_workbook
    wb=load_workbook('test.xlsx')
    sh1=wb['Sheet1']
    v1=sh1.cell(1,1) #获取表格内 （1,1） 坐标位置的值
    print(v1.value)
    v2=sh1['A2'] #获取表格内 A2 坐标位置的值
    print(v2.value)
    v3s=sh1['A1:B2'] #获取表格内 A2-B2 坐标位置的值
    for i in range(len(v3s)):  #用遍历方式将 A2-B2 的值输出
        for a in range(len(v3s[i])):
            print(v3s[i][a].value)

def get_all_value():
    from openpyxl import  load_workbook
    wb = load_workbook('test.xlsx')
    sh = wb['Sheet1']
    for row in sh.rows:   #获取行数
        for cell in row:   #获取每行内的列
            print(cell.value)

    print(sh.max_column)   #获取所有列数
    print(sh.max_row)   #获取所有行数
    print(sh.dimensions)   #获取坐标范围区间 例如A1:C3
# open()
# show_sheets()
# get_one_value()
# get_all_value()